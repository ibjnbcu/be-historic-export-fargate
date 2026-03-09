#!/usr/bin/env python3
import sys
from openpyxl import load_workbook, Workbook
import glob
import os

if len(sys.argv) != 3:
    print("Usage: combine_xlsx.py input_dir output.xlsx", file=sys.stderr)
    sys.exit(1)

input_dir = sys.argv[1]
output_file = sys.argv[2]

files = glob.glob(os.path.join(input_dir, "*.xlsx"))

if not files:
    print("No XLSX files found", file=sys.stderr)
    sys.exit(1)

wb_out = Workbook()
ws_out = wb_out.active

header_written = False

for f in files:
    wb = load_workbook(f)
    ws = wb.active

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            if not header_written:
                ws_out.append(row)
                header_written = True
        else:
            ws_out.append(row)

wb_out.save(output_file)